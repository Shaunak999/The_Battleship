"""
Round-robin AI-vs-AI tournament.

Every AI strategy plays a head-to-head Battleship match against every other
strategy. In each match the two sides defend independently generated random
boards (the same mixed placement style used by the test suite) and take turns
firing at each other, exactly like the engine's "ai_vs_ai" mode. The first
side to sink the opponent's whole fleet wins. Starting player alternates
between games so the first-mover advantage cancels out.

Usage (from the backend/ directory):

    python -m ai.tournament                      # 500 games per pair
    python -m ai.tournament --games 200          # fewer games
    python -m ai.tournament --output my.xlsx     # custom output file

Output
------
Writes a formatted Excel workbook (.xlsx) to --output
(default ai_vs_ai_results.xlsx) with four sheets:

    Summary        - one row per AI: games played, won, lost, draws and the
                     overall win rate, ranked best first.
    Per AI Detail  - for every AI its record against each individual opponent
                     plus a TOTAL row, so you can read "won X, lost Y (Z%)"
                     for every match-up.
    Matchups       - one row per pair of AIs with the full breakdown
                     (wins, losses, draws, win rates, average shots fired).
    Head to Head   - classic matrix: the row AI vs the column AI, shown both
                     as number of games won and as a win percentage.

Plus an interactive HTML dashboard (--report, default ai_vs_ai_results.html)
with head-to-head heatmaps, an overall ranking, matchup breakdowns and the
per-game shot distributions. Requires plotly (see requirements.txt); skip
it with --no-report.

A won/lost breakdown with win percentages is also printed to the console.
"""

import argparse
import os
import sys
import time
from typing import Callable, Dict, List, Optional, Tuple

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Make sure backend/ is importable when run as a plain script too
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from ai.hunt_target_ai import HuntTargetAI
from ai.neural_ai import NeuralAgent
from ai.neural.evaluate import generate_random_board
from ai.probability_ai import ProbabilityAI
from ai.random_ai import RandomAI

# Ordered by strength, used for display and the round-robin pair list.
STRATEGIES: Dict[str, Callable[[], object]] = {
    "random": RandomAI,
    "hunt_target": HuntTargetAI,
    "probability": ProbabilityAI,
    "neural": NeuralAgent,
}

DISPLAY_NAMES = {
    "random": "Random",
    "hunt_target": "Hunt&Target",
    "probability": "Probability",
    "neural": "Neural",
}

MAX_SHOTS_PER_PLAYER = 100  # same cap as the rest of the codebase
MAX_TURNS = 2 * MAX_SHOTS_PER_PLAYER + 20  # hard stop so a game always ends

_SIZE_BY_NAME = {
    "Carrier": 5,
    "Battleship": 4,
    "Cruiser": 3,
    "Submarine": 3,
    "Destroyer": 2,
}


# ----------------------------------------------------------------------
# Excel styling helpers
# ----------------------------------------------------------------------

_HDR_FILL = PatternFill("solid", fgColor="1F4E78")
_HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
_TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
_ALT_FILL = PatternFill("solid", fgColor="F5F8FC")
_DIAG_FILL = PatternFill("solid", fgColor="E2E2E2")
_THIN = Side(style="thin", color="C9C9C9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center")
_PCT = "0.0%"
_AVG = "0.00"


def _style_header(ws, row: int, headers: List[str], start_col: int = 1) -> None:
    """Write a bold, coloured header row with borders."""
    for off, text in enumerate(headers):
        c = ws.cell(row=row, column=start_col + off, value=text)
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.border = _BORDER
        c.alignment = _CENTER


def _autosize(ws, min_width: int = 9, max_width: int = 40) -> None:
    """Size every column to fit its content (approx. width per character)."""
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        longest = max(
            len(str(c.value)) if c.value is not None else 0 for c in col_cells
        )
        # Account for wrapped header text a bit more generously.
        width = max(min_width, min(max_width, longest + 3))
        ws.column_dimensions[letter].width = width


# ----------------------------------------------------------------------
# Head-to-head match
# ----------------------------------------------------------------------

def _fire(
    attacker,
    defender_board: Dict[Tuple[int, int], str],
    ship_hp: Dict[str, set],
    ship_cells_left: int,
) -> Tuple[str, int]:
    """One AI fires at the defender's board.

    Returns (outcome, ship_cells_left_after) where outcome is
    "miss" / "hit" / "sunk" / "win" (the last one sinks the final ship).
    """
    row, col = attacker.choose_move()
    cell = (row, col)

    if cell not in defender_board:
        attacker.process_result(row, col, "miss")
        return "miss", ship_cells_left

    name = defender_board[cell]
    ship_hp[name].discard(cell)
    ship_cells_left -= 1

    if ship_cells_left == 0:
        attacker.process_result(row, col, "sunk", name, _SIZE_BY_NAME[name])
        return "win", 0

    if not ship_hp[name]:
        attacker.process_result(row, col, "sunk", name, _SIZE_BY_NAME[name])
    else:
        attacker.process_result(row, col, "hit")
    return "hit", ship_cells_left


def play_match(
    make_a: Callable[[], object],
    make_b: Callable[[], object],
    a_starts: bool,
) -> Tuple[Optional[str], int, int]:
    """Play one head-to-head match.

    Player A defends ``board_a`` (attacked by B) and attacks ``board_b``.
    Returns (winner, shots_a, shots_b) where winner is "a", "b" or None (draw).
    """
    board_a, place_a = generate_random_board()
    board_b, place_b = generate_random_board()

    ai_a = make_a()
    ai_b = make_b()

    hp_a = {p["name"]: set(tuple(c) for c in p["coordinates"]) for p in place_a}
    hp_b = {p["name"]: set(tuple(c) for c in p["coordinates"]) for p in place_b}
    cells_a = sum(len(v) for v in hp_a.values())  # 17
    cells_b = sum(len(v) for v in hp_b.values())

    shots_a = shots_b = 0

    for turn in range(MAX_TURNS):
        a_fires = (turn % 2 == 0) == a_starts

        if a_fires:
            shots_a += 1
            outcome, cells_b = _fire(ai_a, board_b, hp_b, cells_b)
            if outcome == "win":
                return "a", shots_a, shots_b
        else:
            shots_b += 1
            outcome, cells_a = _fire(ai_b, board_a, hp_a, cells_a)
            if outcome == "win":
                return "b", shots_a, shots_b

        if shots_a >= MAX_SHOTS_PER_PLAYER and shots_b >= MAX_SHOTS_PER_PLAYER:
            break

    return None, shots_a, shots_b  # draw (nobody sank the fleet in time)


# ----------------------------------------------------------------------
# Tournament
# ----------------------------------------------------------------------

def run_tournament(games_per_pair: int) -> Dict[str, Dict]:
    """Run every unordered pair of strategies.

    Returns results[a][b] = dict with keys a_wins, b_wins, draws,
    shots_a (list), shots_b (list) for the a-vs-b session.
    """
    keys = list(STRATEGIES.keys())
    results: Dict[str, Dict[str, Dict]] = {
        a: {b: {"a_wins": 0, "b_wins": 0, "draws": 0, "shots_a": [], "shots_b": [], "winners_a": []} for b in keys}
        for a in keys
    }

    for i in range(len(keys)):
        for j in range(i + 1, len(keys)):
            a, b = keys[i], keys[j]
            stats = results[a][b]
            t0 = time.time()

            for g in range(games_per_pair):
                # Alternate the starter so the first-mover advantage is balanced.
                winner, shots_a, shots_b = play_match(
                    STRATEGIES[a], STRATEGIES[b], a_starts=(g % 2 == 0)
                )
                if winner == "a":
                    stats["a_wins"] += 1
                    stats["winners_a"].append(1)
                elif winner == "b":
                    stats["b_wins"] += 1
                    stats["winners_a"].append(0)
                else:
                    stats["draws"] += 1
                    stats["winners_a"].append(0)
                stats["shots_a"].append(shots_a)
                stats["shots_b"].append(shots_b)

            # Mirror for the reverse direction: b vs a is the complement.
            mirrored = results[b][a]
            mirrored["a_wins"], mirrored["b_wins"] = stats["b_wins"], stats["a_wins"]
            mirrored["draws"] = stats["draws"]
            mirrored["shots_a"] = stats["shots_b"]
            mirrored["shots_b"] = stats["shots_a"]
            mirrored["winners_a"] = [1 if w == 0 and s != None else 0 for w, s in zip(stats["winners_a"], stats["shots_b"])]

            print(
                f"  {DISPLAY_NAMES[a]:<12s} vs {DISPLAY_NAMES[b]:<12s} "
                f"({games_per_pair} games, {time.time() - t0:.0f}s) "
                f"-> {DISPLAY_NAMES[a]} {stats['a_wins']} / "
                f"{DISPLAY_NAMES[b]} {stats['b_wins']} / draws {stats['draws']}",
                flush=True,
            )

    return results


# ----------------------------------------------------------------------
# Console report
# ----------------------------------------------------------------------

def print_records(results: Dict[str, Dict], games_per_pair: int) -> None:
    """Print how many games each AI won and lost, with win percentages.

    For every AI it shows the record against each opponent plus the overall
    record across all its games.
    """
    keys = list(STRATEGIES.keys())
    per_ai = []  # (total_wins, total_losses, name, lines, draw_txt, overall_pct)
    for a in keys:
        lines = []
        total_wins = total_losses = total_draws = 0
        for b in keys:
            if a == b:
                continue
            stats = results[a][b]
            wins, losses = stats["a_wins"], stats["b_wins"]
            draws = stats["draws"]
            total_wins += wins
            total_losses += losses
            total_draws += draws
            decided = wins + losses
            pct = 100.0 * wins / max(decided, 1)
            draw_txt = f", drawn {draws}" if draws else ""
            lines.append(
                f"    vs {DISPLAY_NAMES[b]:<12s}: "
                f"won {wins:>4}  lost {losses:>4}{draw_txt}  "
                f"({pct:5.1f}%)"
            )
        total = games_per_pair * (len(keys) - 1)
        decided = total - total_draws
        overall_pct = 100.0 * total_wins / max(decided, 1)
        draw_txt = f", drawn {total_draws}" if total_draws else ""
        per_ai.append((total_wins, total_losses, a, lines, draw_txt, overall_pct))

    print(f"\n  Won/lost record per AI ({games_per_pair} games per match-up)\n")
    for total_wins, total_losses, a, lines, draw_txt, overall_pct in sorted(
        per_ai, key=lambda x: -x[0]
    ):
        print(f"  {DISPLAY_NAMES[a]}:")
        for line in lines:
            print(line)
        print(
            f"    {'TOTAL':<12s}: won {total_wins:>4}  lost {total_losses:>4}"
            f"{draw_txt}  ({overall_pct:5.1f}%)"
        )
        print()


# ----------------------------------------------------------------------
# Excel report
# ----------------------------------------------------------------------

def _ai_totals(results: Dict[str, Dict], games_per_pair: int) -> Dict[str, Tuple]:
    """Per AI: (games, won, lost, draws) across all its match-ups."""
    keys = list(STRATEGIES.keys())
    totals = {}
    for a in keys:
        w = l = d = 0
        for b in keys:
            if a == b:
                continue
            s = results[a][b]
            w += s["a_wins"]
            l += s["b_wins"]
            d += s["draws"]
        totals[a] = (games_per_pair * (len(keys) - 1), w, l, d)
    return totals


def _build_summary_sheet(wb, results, games_per_pair) -> None:
    ws = wb.active
    ws.title = "Summary"

    headers = ["Rank", "AI Strategy", "Games Played", "Won", "Lost", "Draws", "Win Rate"]
    _style_header(ws, 1, headers)

    totals = _ai_totals(results, games_per_pair)
    rows = sorted(totals.items(), key=lambda kv: -kv[1][1])  # most wins first

    for rank, (a, (games, w, l, d)) in enumerate(rows, start=1):
        r = rank + 1
        ws.cell(row=r, column=1, value=rank).alignment = _CENTER
        ws.cell(row=r, column=2, value=DISPLAY_NAMES[a])
        ws.cell(row=r, column=3, value=games).alignment = _CENTER
        ws.cell(row=r, column=4, value=w).alignment = _CENTER
        ws.cell(row=r, column=5, value=l).alignment = _CENTER
        ws.cell(row=r, column=6, value=d).alignment = _CENTER
        pct = ws.cell(row=r, column=7, value=w / max(games - d, 1))
        pct.number_format = _PCT
        pct.alignment = _CENTER
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).border = _BORDER
            if rank % 2 == 0:
                ws.cell(row=r, column=col).fill = _ALT_FILL

    _autosize(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _build_per_ai_sheet(wb, results, games_per_pair) -> None:
    ws = wb.create_sheet("Per AI Detail")
    headers = ["AI Strategy", "Opponent", "Games", "Won", "Lost", "Draws", "Win Rate"]
    _style_header(ws, 1, headers)

    keys = list(STRATEGIES.keys())
    row = 2
    for ai_idx, a in enumerate(keys):
        block_start = row
        for b in keys:
            if a == b:
                continue
            s = results[a][b]
            w, l, d = s["a_wins"], s["b_wins"], s["draws"]
            ws.cell(row=row, column=1, value=DISPLAY_NAMES[a])
            ws.cell(row=row, column=2, value=f"vs {DISPLAY_NAMES[b]}")
            ws.cell(row=row, column=3, value=games_per_pair)
            ws.cell(row=row, column=4, value=w)
            ws.cell(row=row, column=5, value=l)
            ws.cell(row=row, column=6, value=d)
            pct = ws.cell(row=row, column=7, value=w / max(w + l, 1))
            pct.number_format = _PCT
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=row, column=col)
                c.border = _BORDER
                c.alignment = _LEFT if col <= 2 else _CENTER
            row += 1

        # TOTAL row for this AI
        total_row = row
        games, w, l, d = _ai_totals(results, games_per_pair)[a]
        ws.cell(row=total_row, column=1, value=DISPLAY_NAMES[a])
        ws.cell(row=total_row, column=2, value="TOTAL")
        ws.cell(row=total_row, column=3, value=games)
        ws.cell(row=total_row, column=4, value=w)
        ws.cell(row=total_row, column=5, value=l)
        ws.cell(row=total_row, column=6, value=d)
        pct = ws.cell(row=total_row, column=7, value=w / max(games - d, 1))
        pct.number_format = _PCT
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=total_row, column=col)
            c.fill = _TOTAL_FILL
            c.font = Font(bold=True)
            c.border = _BORDER
            c.alignment = _LEFT if col <= 2 else _CENTER
        row = total_row + 1  # blank spacer row between AI blocks

        # Light banding for the whole AI block (except the TOTAL row)
        if ai_idx % 2 == 1:
            for r in range(block_start, total_row):
                for col in range(1, len(headers) + 1):
                    ws.cell(row=r, column=col).fill = _ALT_FILL

    _autosize(ws)
    ws.freeze_panes = "A2"


def _build_matchups_sheet(wb, results, games_per_pair) -> None:
    ws = wb.create_sheet("Matchups")
    headers = [
        "Matchup", "AI A", "AI B", "Games Played",
        "AI A Won", "AI B Won", "Draws",
        "AI A Win Rate", "AI B Win Rate",
        "Avg Shots AI A", "Avg Shots AI B",
    ]
    _style_header(ws, 1, headers)

    keys = list(STRATEGIES.keys())
    row = 2
    for i in range(len(keys)):
        for j in range(i + 1, len(keys)):
            a, b = keys[i], keys[j]
            s = results[a][b]
            shots_a = np.mean(s["shots_a"])
            shots_b = np.mean(s["shots_b"])

            ws.cell(row=row, column=1, value=f"{DISPLAY_NAMES[a]} vs {DISPLAY_NAMES[b]}")
            ws.cell(row=row, column=2, value=DISPLAY_NAMES[a])
            ws.cell(row=row, column=3, value=DISPLAY_NAMES[b])
            ws.cell(row=row, column=4, value=games_per_pair)
            ws.cell(row=row, column=5, value=s["a_wins"])
            ws.cell(row=row, column=6, value=s["b_wins"])
            ws.cell(row=row, column=7, value=s["draws"])
            for col, val in ((8, s["a_wins"] / games_per_pair),
                             (9, s["b_wins"] / games_per_pair)):
                c = ws.cell(row=row, column=col, value=val)
                c.number_format = _PCT
            for col, val in ((10, shots_a), (11, shots_b)):
                c = ws.cell(row=row, column=col, value=val)
                c.number_format = _AVG

            for col in range(1, len(headers) + 1):
                c = ws.cell(row=row, column=col)
                c.border = _BORDER
                if col != 1:
                    c.alignment = _CENTER
                if row % 2 == 1:
                    c.fill = _ALT_FILL
            row += 1

    _autosize(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _build_head_to_head_sheet(wb, results, games_per_pair) -> None:
    ws = wb.create_sheet("Head to Head")
    keys = list(STRATEGIES.keys())
    n = len(keys)

    # Block 1: games won (row AI beats column AI)
    title1 = ws.cell(row=1, column=1, value="Games WON — the row AI beats the column AI")
    title1.font = Font(bold=True, size=12)
    _style_header(ws, 2, [""] + [DISPLAY_NAMES[k] for k in keys])
    for i, a in enumerate(keys):
        r = 3 + i
        ws.cell(row=r, column=1, value=DISPLAY_NAMES[a]).font = Font(bold=True)
        for j, b in enumerate(keys):
            c = ws.cell(row=r, column=2 + j)
            c.border = _BORDER
            c.alignment = _CENTER
            if a == b:
                c.value = "—"
                c.fill = _DIAG_FILL
            else:
                c.value = results[a][b]["a_wins"]
                if i % 2 == 1:
                    c.fill = _ALT_FILL

    # Block 2: win percentage, same layout, one blank column apart
    start2 = n + 3
    title2 = ws.cell(row=1, column=start2, value="WIN RATE — the row AI beats the column AI")
    title2.font = Font(bold=True, size=12)
    _style_header(ws, 2, [""] + [DISPLAY_NAMES[k] for k in keys], start_col=start2)
    for i, a in enumerate(keys):
        for j, b in enumerate(keys):
            c = ws.cell(row=3 + i, column=start2 + j)
            c.border = _BORDER
            c.alignment = _CENTER
            if a == b:
                c.value = "—"
                c.fill = _DIAG_FILL
            else:
                s = results[a][b]
                c.value = s["a_wins"] / max(s["a_wins"] + s["b_wins"], 1)
                c.number_format = _PCT
                if i % 2 == 1:
                    c.fill = _ALT_FILL

    note = ws.cell(
        row=n + 4, column=1,
        value="Win rate = games won / (won + lost); drawn games excluded. "
              f"{games_per_pair} games per match-up, starting player alternated.",
    )
    note.font = Font(italic=True, color="595959")

    _autosize(ws)
    ws.freeze_panes = "A3"


def write_excel(
    results: Dict[str, Dict],
    games_per_pair: int,
    path: str,
) -> None:
    """Write the tournament statistics to a formatted Excel workbook."""
    wb = Workbook()
    _build_summary_sheet(wb, results, games_per_pair)
    _build_per_ai_sheet(wb, results, games_per_pair)
    _build_matchups_sheet(wb, results, games_per_pair)
    _build_head_to_head_sheet(wb, results, games_per_pair)
    try:
        wb.save(path)
        print(f"\n  Excel workbook written to {os.path.abspath(path)}")
    except PermissionError:
        print(f"\n  [!] Could not write Excel workbook to {os.path.abspath(path)} (file is open in another program).")



# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Round-robin AI-vs-AI tournament")
    parser.add_argument(
        "--games", type=int, default=500,
        help="Games per match-up (default: 500)",
    )
    parser.add_argument(
        "--output", type=str, default="ai_vs_ai_results.xlsx",
        help="Excel file to write the statistics to "
             "(default: ai_vs_ai_results.xlsx)",
    )
    parser.add_argument(
        "--report", type=str, default="ai_vs_ai_results.html",
        help="Interactive HTML report to write "
             "(default: ai_vs_ai_results.html)",
    )
    parser.add_argument(
        "--no-report", action="store_true",
        help="Skip the interactive HTML report",
    )
    parser.add_argument(
        "--no-open", action="store_true",
        help="Do not automatically open the HTML report in default browser",
    )
    args = parser.parse_args()

    print("=" * 62)
    print("  AI VS AI — ROUND ROBIN TOURNAMENT")
    print("=" * 62)
    print(f"  Strategies : {', '.join(DISPLAY_NAMES[k] for k in STRATEGIES)}")
    print(f"  Games/pair : {args.games}  (starting player alternates)")
    print(f"  Pairs      : {len(STRATEGIES) * (len(STRATEGIES) - 1) // 2}")
    print()

    t0 = time.time()
    results = run_tournament(args.games)
    print(f"\n  Tournament finished in {time.time() - t0:.0f}s")
    print_records(results, args.games)
    write_excel(results, args.games, args.output)
    if not args.no_report:
        try:
            from ai.report import write_report
            write_report(results, args.games, args.report, auto_open=not args.no_open)
        except ImportError as exc:
            print(
                f"\n  [!] HTML report skipped: plotly is not installed ({exc}).\n"
                "      Add 'plotly' to requirements.txt and reinstall, "
                "or rerun with --no-report."
            )


if __name__ == "__main__":
    main()
